import csv
from openpyxl import Workbook, load_workbook
from loguru import logger
import pkg_resources
from symspellpy import SymSpell
import sys


class Transform:
    """
    Transform contains methods for converting
        1. csv  --> xlsx
        2. xlsx --> csv
    """

    def __init__(self, verbose: bool = False, log: str = "seiri-error.log") -> None:
        # Initialize logger for default logging
        self.logger = logger
        self.logger.remove()
        self.logger.add(
            sink=log,
            level="ERROR",
            format="<white>{time:MMMM D, YYYY > HH:mm:ss}</white> | <level>{level: <8}</level> | <level>Transform</level> | <level>{message}</level>",
        )

        if verbose:
            self.logger.add(
                sink=sys.stdout,
                level="DEBUG",
                format="<white>{time:MMMM D, YYYY > HH:mm:ss}</white> | <level>{level: <8}</level> | <level>Transform</level> | <level>{message}</level>",
            )

        # Intialize worksheets
        self.wb = Workbook()
        self.sheets = []
        del self.wb["Sheet"]

        # default langs for now
        self.langs = ["en", "de", "es", "fr", "it"]

        # Intialize spell checker
        self.sym_spell = SymSpell(max_dictionary_edit_distance=2, prefix_length=7)
        dictionary_path = pkg_resources.resource_filename(
            "symspellpy", "frequency_dictionary_en_82_765.txt"
        )
        bigram_path = pkg_resources.resource_filename(
            "symspellpy", "frequency_bigramdictionary_en_243_342.txt"
        )

        self.sym_spell.load_dictionary(dictionary_path, term_index=0, count_index=1)
        self.sym_spell.load_bigram_dictionary(bigram_path, term_index=0, count_index=2)

    def __init__sheets(self) -> None:
        # Create worksheets
        for lang in self.langs:
            self.sheets.append(self.wb.create_sheet(lang))
            self.logger.info(f"Created sheet {lang}")

        for sheet in self.sheets:
            sheet.append(["Key", "Value"])

        self.logger.success("Successfully initialized Sheets")

    def csv_to_xlsx(self, in_file: str, out_file: str) -> None:
        Listed_csv = csv.reader(open(in_file, "r", newline=""), delimiter=";")

        Listed_csv.__next__()

        self.__init__sheets()
        self.logger.info("Spell checking column 'en'")
        for row in Listed_csv:
            if len(row):  # Don't consider empty lines
                ## Spell check
                suggestions = self.sym_spell.lookup_compound(
                    row[4], max_edit_distance=2, transfer_casing=True
                )
                if suggestions:
                    if suggestions[0].term != row[4]:
                        self.logger.error(
                            f"Spelling mistake found in {row[4]} : Suggestions: {suggestions[0].term}"
                        )

                self.sheets[0].append([row[2], row[4]])
        self.logger.success("Spell checked column 'en'")
        # Save Workbook
        self.wb.save(out_file)
        self.logger.info(f"Saving workbook to {out_file}")

    def xlsx_to_csv(self, in_file: str, out_file: str) -> None:
        # initialize workbook
        wb = load_workbook(in_file)
        wb.sheetnames
        self.logger.success(f"Loaded {in_file}")

        # initialize csv writer
        csv_file = open(out_file, "w", newline="")
        csv_writer = csv.writer(csv_file, delimiter=";")
        self.logger.info(f"Writing to {out_file}")

        # write headers "Section Number;Entry Number;Reference;(default);en;de;es;fr;it;Review ". Some of the headers are from self.langs so iterate over them as well
        headers = ["Section Number", "Entry Number", "Reference", "(default)"]
        headers.extend(self.langs)
        headers.append("Review")
        csv_writer.writerow(headers)
        self.logger.success(f"Headers written to {out_file}")

        # Write rows to csv
        # each row looks like this
        # 0;0;lbl_m_access_level;;Access Level;sdfk;ryt;df;g;

        for row in range(2, wb["en"].max_row + 1):
            row_to_append = [
                0,
                0,
                wb["en"].cell(row=row, column=1).value,
                "",
            ]
            for sheet in wb.sheetnames:
                row_to_append.append(wb[sheet].cell(row=row, column=2).value)
            row_to_append.append("")

            csv_writer.writerow(row_to_append)
        self.logger.success(f"Rows written to {out_file}")

        # close csv file
        csv_file.close()
        self.logger.success(f"Successfully converted {in_file} to {out_file}")


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--cx", type=str, help="Convert csv to xlsx")
    ap.add_argument("-o", "--output", required=False, type=str, help="output file")
    ap.add_argument("--xc", type=str, help="Convert xlsx to csv")

    ap.add_argument(
        "--log",
        type=str,
        default="seiri-error.log",
        help="path to log file",
    )
    ap.add_argument("-v", "--verbose", action="store_true")

    args = vars(ap.parse_args())

    transformer = Transform(verbose=args["verbose"], log=args["log"])

    if args["cx"]:
        # csv to xlsx
        output = args["cx"][:-3] + "xlsx" if not args["output"] else args["output"]
        input = args["cx"]

        transformer.csv_to_xlsx(input, output)

    if args["xc"]:
        # xlsx to csv
        output = args["xc"][:-4] + "csv" if not args["output"] else args["output"]
        input = args["xc"]

        transformer.xlsx_to_csv(input, output)
